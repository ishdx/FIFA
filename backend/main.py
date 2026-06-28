from fastapi import FastAPI, HTTPException, Depends, status, Request, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from contextlib import asynccontextmanager
from pydantic import BaseModel
from typing import Optional
import pg8000.native, json, os, secrets, asyncio, urllib.parse, traceback, io
from datetime import datetime

# ── SSE broadcaster ───────────────────────────────────────
class EventBroadcaster:
    def __init__(self):
        self._queues = []
    def subscribe(self):
        q = asyncio.Queue(maxsize=10)
        self._queues.append(q)
        return q
    def unsubscribe(self, q):
        if q in self._queues:
            self._queues.remove(q)
    async def broadcast(self, event: dict):
        dead = []
        for q in self._queues:
            try: q.put_nowait(event)
            except asyncio.QueueFull: dead.append(q)
        for q in dead:
            if q in self._queues: self._queues.remove(q)

broadcaster = EventBroadcaster()

@asynccontextmanager
async def lifespan(app: FastAPI):
    asyncio.create_task(asyncio.to_thread(startup))
    yield

app = FastAPI(title="FIFA WC2026 Dashboard API", lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

DATABASE_URL = os.environ.get("DATABASE_URL", "")
ADMIN_USER   = os.environ.get("ADMIN_USER", "admin")
ADMIN_PASS   = os.environ.get("ADMIN_PASS", "smi2026")

# ── DB ────────────────────────────────────────────────────
def get_conn():
    r = urllib.parse.urlparse(DATABASE_URL)
    return pg8000.native.Connection(
        host=r.hostname,
        port=r.port or 5432,
        database=r.path.lstrip("/"),
        user=urllib.parse.unquote(r.username or ""),
        password=urllib.parse.unquote(r.password or ""),
        ssl_context=True,
    )

def _esc(val):
    """Escape a Python value for safe interpolation into a PostgreSQL query."""
    if val is None:
        return "NULL"
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        return repr(val)
    # str / anything else: escape single quotes for PostgreSQL standard strings
    return "'" + str(val).replace("'", "''") + "'"

def db_run(conn, sql, *args):
    """Execute SQL using pg8000 execute_simple (no extended protocol / bind params).
    Use %s as placeholder; values are safely escaped and interpolated."""
    if args:
        parts = sql.split("%s")
        sql = "".join(p + _esc(v) for p, v in zip(parts, args)) + parts[-1]
    return conn.run(sql)   # len(params)==0 and stream is None → execute_simple

def safe_run(conn, sql, *args):
    try:
        return db_run(conn, sql, *args)
    except Exception as e:
        print(f"SQL error: {e}\nSQL: {sql}")
        return []

# ── Auth ──────────────────────────────────────────────────
def require_admin(x_admin_token: str = Header(default="")):
    import base64
    try:
        decoded = base64.b64decode(x_admin_token).decode()
        user, password = decoded.split(":", 1)
        if secrets.compare_digest(user, ADMIN_USER) and secrets.compare_digest(password, ADMIN_PASS):
            return user
    except Exception:
        pass
    raise HTTPException(status_code=401, detail="Unauthorized")

# ── Scoring ───────────────────────────────────────────────
def check_pred(prediction: str, result: str) -> bool:
    if not prediction or not result:
        return False
    p_en = prediction.lower().split("|")[-1].strip()
    r_en = result.lower().split("|")[-1].strip()
    return p_en == r_en

def recalc_all_points(conn):
    results_rows = safe_run(conn,
        "SELECT round, match_name, result FROM matches WHERE status='done' AND result IS NOT NULL")
    results = {(r[0], r[1]): r[2] for r in results_rows}
    all_ids = safe_run(conn, "SELECT emp_id FROM participants")
    for row in all_ids:
        emp_id = row[0]
        preds = safe_run(conn,
            "SELECT round, match_name, prediction FROM predictions WHERE emp_id=%s", emp_id)
        pts = {1: 0.0, 2: 0.0, 3: 0.0, 4: 0.0}
        for pr in preds:
            key = (pr[0], pr[1])
            if key in results and pr[2] and check_pred(pr[2], results[key]):
                pts[pr[0]] += 3.0
        bonus_row = safe_run(conn, "SELECT bonus_pts FROM points_cache WHERE emp_id=%s", emp_id)
        bonus_pts = float(bonus_row[0][0]) if bonus_row else 0.0
        total = pts[1] + pts[2] + pts[3] + pts[4] + bonus_pts
        safe_run(conn, """INSERT INTO points_cache (emp_id,r1_pts,r2_pts,r3_pts,bonus_pts,total)
                    VALUES (%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (emp_id) DO UPDATE SET
                    r1_pts=EXCLUDED.r1_pts, r2_pts=EXCLUDED.r2_pts,
                    r3_pts=EXCLUDED.r3_pts, total=EXCLUDED.total""",
                 emp_id, pts[1], pts[2], pts[3], bonus_pts, total)

# ── DB init ───────────────────────────────────────────────
def init_db(conn):
    safe_run(conn, """CREATE TABLE IF NOT EXISTS participants (
        emp_id TEXT PRIMARY KEY, name TEXT NOT NULL, rounds TEXT DEFAULT '[]')""")
    safe_run(conn, """CREATE TABLE IF NOT EXISTS predictions (
        emp_id TEXT, round INTEGER, match_name TEXT, prediction TEXT,
        PRIMARY KEY (emp_id, round, match_name))""")
    safe_run(conn, """CREATE TABLE IF NOT EXISTS matches (
        id SERIAL PRIMARY KEY, round INTEGER NOT NULL, match_name TEXT NOT NULL,
        result TEXT, status TEXT DEFAULT 'pending', played_at TEXT,
        UNIQUE(round, match_name))""")
    safe_run(conn, """CREATE TABLE IF NOT EXISTS points_cache (
        emp_id TEXT PRIMARY KEY,
        r1_pts REAL DEFAULT 0, r2_pts REAL DEFAULT 0,
        r3_pts REAL DEFAULT 0, total REAL DEFAULT 0)""")
    safe_run(conn, "ALTER TABLE points_cache ADD COLUMN IF NOT EXISTS bonus_pts REAL DEFAULT 0")
    safe_run(conn, "ALTER TABLE matches ADD COLUMN IF NOT EXISTS options TEXT DEFAULT '[]'")

# ── Seed ─────────────────────────────────────────────────
def _batch_insert(conn, table, cols, rows_of_vals, batch=200):
    """INSERT rows in batches using multi-row VALUES — one round-trip per batch."""
    col_str = f"({','.join(cols)})"
    for i in range(0, len(rows_of_vals), batch):
        chunk = rows_of_vals[i:i+batch]
        vals  = ",".join(
            "(" + ",".join(_esc(v) for v in row) + ")"
            for row in chunk
        )
        db_run(conn, f"INSERT INTO {table} {col_str} VALUES {vals} ON CONFLICT DO NOTHING")

def seed_db(conn):
    count = db_run(conn, "SELECT COUNT(*) FROM participants")[0][0]
    if count > 0:
        print(f"Already seeded ({count} participants), skipping.")
        return

    seed_path = os.path.join(os.path.dirname(__file__), "..", "data", "seed_data.json")
    if not os.path.exists(seed_path):
        print("No seed_data.json found.")
        return

    with open(seed_path, encoding="utf-8") as f:
        data = json.load(f)

    match_options = data.get("match_options", {})
    participants  = data["participants"]
    print(f"Seeding {len(participants)} participants...")

    db_run(conn, "BEGIN")
    try:
        _batch_insert(conn, "participants", ["emp_id","name","rounds"],
            [(eid, info["name"], json.dumps(info["rounds"]))
             for eid, info in participants.items()])
        print("  participants done")

        _batch_insert(conn, "points_cache",
            ["emp_id","r1_pts","r2_pts","r3_pts","bonus_pts","total"],
            [(eid, 0, 0, 0, 0, 0) for eid in participants])
        print("  points_cache done")

        preds = []
        for eid, info in participants.items():
            for pred_key, pred_val in info["predictions"].items():
                rnd_str, mname = pred_key.split(":", 1)
                preds.append((eid, int(rnd_str[1]), mname, pred_val or ""))
        _batch_insert(conn, "predictions", ["emp_id","round","match_name","prediction"], preds)
        print(f"  predictions done: {len(preds)}")

        matches = []
        for rnd, key in [(1,"R1"),(2,"R2"),(3,"R3")]:
            for mname in data["matches"][key]:
                opts = json.dumps(match_options.get(f"{key}:{mname}", []), ensure_ascii=False)
                matches.append((rnd, mname, "pending", opts))
        _batch_insert(conn, "matches", ["round","match_name","status","options"], matches)
        print("  matches done")

        db_run(conn, "COMMIT")
        print(f"Seeding complete: {len(participants)} participants, {len(preds)} predictions!")

    except Exception as e:
        db_run(conn, "ROLLBACK")
        print(f"Seed failed, rolled back: {e}")
        raise


def startup():
    try:
        conn = get_conn()
        try:
            init_db(conn)
            seed_db(conn)
        finally:
            conn.close()
    except Exception as e:
        print(f"STARTUP ERROR: {e}")
        traceback.print_exc()

# ── PUBLIC ENDPOINTS ──────────────────────────────────────
@app.get("/api/health")
def health():
    try:
        conn = get_conn()
        conn.run("SELECT 1")
        try:
            count = db_run(conn, "SELECT COUNT(*) FROM participants")[0][0]
            matches = db_run(conn, "SELECT COUNT(*) FROM matches")[0][0]
        except:
            count, matches = 0, 0
        conn.close()
        return {"status": "ok", "db": "connected", "participants": count, "matches": matches}
    except Exception as e:
        return {"status": "error", "detail": str(e)}

@app.get("/api/stats")
def get_stats():
    conn = get_conn()
    try:
        total   = safe_run(conn, "SELECT COUNT(*) FROM participants") or [[0]]
        total   = total[0][0]
        r1c     = (safe_run(conn, "SELECT COUNT(DISTINCT emp_id) FROM predictions WHERE round=1") or [[0]])[0][0]
        r2c     = (safe_run(conn, "SELECT COUNT(DISTINCT emp_id) FROM predictions WHERE round=2") or [[0]])[0][0]
        r3c     = (safe_run(conn, "SELECT COUNT(DISTINCT emp_id) FROM predictions WHERE round=3") or [[0]])[0][0]
        r4c     = (safe_run(conn, "SELECT COUNT(DISTINCT emp_id) FROM predictions WHERE round=4") or [[0]])[0][0]
        all3    = (safe_run(conn, "SELECT COUNT(*) FROM (SELECT emp_id FROM predictions GROUP BY emp_id HAVING COUNT(DISTINCT round)=3) x") or [[0]])[0][0]
        top     = safe_run(conn, "SELECT emp_id,r1_pts,r2_pts,r3_pts,COALESCE(bonus_pts,0),total FROM points_cache ORDER BY total DESC LIMIT 1") or []
        avg_row = safe_run(conn, "SELECT AVG(total),AVG(r1_pts),AVG(r2_pts),AVG(r3_pts) FROM points_cache") or [[0,0,0,0]]
        avg     = avg_row[0]
        mn      = (safe_run(conn, "SELECT MIN(total) FROM points_cache") or [[0]])[0][0]
        done    = (safe_run(conn, "SELECT COUNT(*) FROM matches WHERE status='done'") or [[0]])[0][0]
        pending = (safe_run(conn, "SELECT COUNT(*) FROM matches WHERE status='pending'") or [[0]])[0][0]

        top_data = {"name":"","total":0,"r1":0,"r2":0,"r3":0,"bonus":0}
        if top:
            t = top[0]
            p = safe_run(conn, "SELECT name FROM participants WHERE emp_id=%s", t[0])
            top_data = {"name":p[0][0] if p else "","total":float(t[5] or 0),
                        "r1":float(t[1] or 0),"r2":float(t[2] or 0),"r3":float(t[3] or 0),"bonus":float(t[4] or 0)}

        combos = {}
        for k, q in [
            ("r1only","SELECT COUNT(DISTINCT emp_id) FROM predictions WHERE round=1 AND emp_id NOT IN (SELECT emp_id FROM predictions WHERE round=2) AND emp_id NOT IN (SELECT emp_id FROM predictions WHERE round=3)"),
            ("r2only","SELECT COUNT(DISTINCT emp_id) FROM predictions WHERE round=2 AND emp_id NOT IN (SELECT emp_id FROM predictions WHERE round=1) AND emp_id NOT IN (SELECT emp_id FROM predictions WHERE round=3)"),
            ("r3only","SELECT COUNT(DISTINCT emp_id) FROM predictions WHERE round=3 AND emp_id NOT IN (SELECT emp_id FROM predictions WHERE round=1) AND emp_id NOT IN (SELECT emp_id FROM predictions WHERE round=2)"),
            ("r1r2","SELECT COUNT(DISTINCT p1.emp_id) FROM predictions p1 JOIN predictions p2 ON p1.emp_id=p2.emp_id WHERE p1.round=1 AND p2.round=2 AND p1.emp_id NOT IN (SELECT emp_id FROM predictions WHERE round=3)"),
            ("r1r3","SELECT COUNT(DISTINCT p1.emp_id) FROM predictions p1 JOIN predictions p2 ON p1.emp_id=p2.emp_id WHERE p1.round=1 AND p2.round=3 AND p1.emp_id NOT IN (SELECT emp_id FROM predictions WHERE round=2)"),
            ("r2r3","SELECT COUNT(DISTINCT p1.emp_id) FROM predictions p1 JOIN predictions p2 ON p1.emp_id=p2.emp_id WHERE p1.round=2 AND p2.round=3 AND p1.emp_id NOT IN (SELECT emp_id FROM predictions WHERE round=1)"),
        ]:
            combos[k] = db_run(conn, q)[0][0]
        combos["all3"] = all3

        totals = [r[0] for r in db_run(conn, "SELECT total FROM points_cache")]
        bins = [0,30,40,50,60,70,80,90,100,110,120,130,500]
        bl   = ['0-30','31-40','41-50','51-60','61-70','71-80','81-90','91-100','101-110','111-120','121-130','130+']
        dist = [0]*len(bl)
        for t in totals:
            for i in range(len(bins)-1):
                if (i==0 and t==0) or (bins[i]<t<=bins[i+1]):
                    dist[i]+=1; break

        return {"total_participants":total,"r1_participants":r1c,"r2_participants":r2c,"r3_participants":r3c,
                "all3_participants":all3,"leader":top_data,
                "avg_total":round(float(avg[0] or 0),1),"avg_r1":round(float(avg[1] or 0),1),
                "avg_r2":round(float(avg[2] or 0),1),"avg_r3":round(float(avg[3] or 0),1),
                "min_total":float(mn or 0),"matches_done":done,"matches_pending":pending,
                "r4_participants":r4c,
                "avg_r4":round(float((safe_run(conn,"SELECT AVG(COALESCE(r4_pts,0)) FROM points_cache WHERE emp_id IN (SELECT DISTINCT emp_id FROM predictions WHERE round=4)") or [[0]])[0][0] or 0),1),
                "combos":combos,"dist_labels":bl,"dist_values":dist}
    finally:
        conn.close()

@app.get("/api/leaderboard")
def get_leaderboard(page: int=1, per_page: int=20, search: str="", round_filter: str="all"):
    conn = get_conn()
    try:
        rows = db_run(conn, """SELECT pc.emp_id,p.name,pc.r1_pts,pc.r2_pts,pc.r3_pts,
                           COALESCE(pc.bonus_pts,0),pc.total,p.rounds
                           FROM points_cache pc JOIN participants p ON pc.emp_id=p.emp_id
                           ORDER BY pc.total DESC, p.name ASC""")
        results = []
        for r in rows:
            rounds = json.loads(r[7])
            if round_filter=="r1" and 1 not in rounds: continue
            if round_filter=="r2" and 2 not in rounds: continue
            if round_filter=="r3" and 3 not in rounds: continue
            if round_filter=="r4" and 4 not in rounds: continue
            if round_filter=="all3" and not all(x in rounds for x in [1,2,3]): continue
            if search and search.lower() not in r[1].lower() and search not in r[0]: continue
            results.append({"rank":len(results)+1,"emp_id":r[0],"name":r[1],
                            "r1":float(r[2]),"r2":float(r[3]),"r3":float(r[4]),
                            "r4":float(r[5]),"bonus":float(r[6]),"total":float(r[7]),
                            "p1":1 in rounds,"p2":2 in rounds,"p3":3 in rounds,"p4":4 in rounds})
        start=(page-1)*per_page
        return {"data":results[start:start+per_page],"total":len(results),"page":page,"per_page":per_page}
    finally:
        conn.close()

@app.get("/api/matches")
def get_matches():
    conn = get_conn()
    try:
        rows = db_run(conn, "SELECT id,round,match_name,result,status,played_at,COALESCE(options,'[]') FROM matches ORDER BY round,id")
        return [{"id":r[0],"round":r[1],"match_name":r[2],"result":r[3],
                 "status":r[4],"played_at":r[5],"options":json.loads(r[6])} for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        conn.close()

@app.get("/api/events")
async def sse_events(request: Request):
    q = broadcaster.subscribe()
    async def event_stream():
        yield "data: {\"type\":\"connected\"}\n\n"
        try:
            while True:
                if await request.is_disconnected(): break
                try:
                    event = await asyncio.wait_for(q.get(), timeout=25)
                    yield f"data: {json.dumps(event)}\n\n"
                except asyncio.TimeoutError:
                    yield "data: {\"type\":\"ping\"}\n\n"
        finally:
            broadcaster.unsubscribe(q)
    return StreamingResponse(event_stream(), media_type="text/event-stream",
                             headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

# ── ADMIN ENDPOINTS ───────────────────────────────────────
@app.get("/api/admin/verify")
def verify_admin(_=Depends(require_admin)):
    return {"status": "ok"}

class MatchResult(BaseModel):
    round: int
    match_name: str
    result: str

@app.post("/api/admin/result")
async def submit_result(payload: MatchResult, _=Depends(require_admin)):
    conn = get_conn()
    try:
        ex = db_run(conn, "SELECT id FROM matches WHERE round=%s AND match_name=%s",
                    payload.round, payload.match_name)
        if not ex:
            raise HTTPException(404, "Match not found")
        db_run(conn, "UPDATE matches SET result=%s,status='done',played_at=%s WHERE round=%s AND match_name=%s",
               payload.result, datetime.utcnow().isoformat(), payload.round, payload.match_name)
    finally:
        conn.close()
    await broadcaster.broadcast({"type":"match_saved","match":payload.match_name,
                                  "result":payload.result,"round":payload.round})
    return {"status":"ok","match":payload.match_name,"result":payload.result}

@app.post("/api/admin/recalculate")
async def recalculate(_=Depends(require_admin)):
    conn = get_conn()
    try:
        recalc_all_points(conn)
        count = db_run(conn, "SELECT COUNT(*) FROM participants")[0][0]
    finally:
        conn.close()
    await broadcaster.broadcast({"type":"update"})
    return {"status":"ok","participants_recalculated":count}

class BonusPoints(BaseModel):
    emp_id: str
    bonus_pts: float

@app.post("/api/admin/bonus")
async def set_bonus(payload: BonusPoints, _=Depends(require_admin)):
    conn = get_conn()
    try:
        ex = db_run(conn, "SELECT emp_id FROM participants WHERE emp_id=%s", payload.emp_id)
        if not ex:
            raise HTTPException(404, "Participant not found")
        db_run(conn, """UPDATE points_cache SET bonus_pts=%s,
                    total=r1_pts+r2_pts+r3_pts+%s WHERE emp_id=%s""",
               payload.bonus_pts, payload.bonus_pts, payload.emp_id)
    finally:
        conn.close()
    await broadcaster.broadcast({"type":"update","bonus":True})
    return {"status":"ok","emp_id":payload.emp_id,"bonus_pts":payload.bonus_pts}

@app.get("/api/admin/matches/pending")
def get_pending(_=Depends(require_admin)):
    conn = get_conn()
    try:
        rows = db_run(conn, "SELECT id,round,match_name,result,status,COALESCE(options,'[]') FROM matches WHERE status='pending' ORDER BY round,id")
        return [{"id":r[0],"round":r[1],"match_name":r[2],"result":r[3],
                 "status":r[4],"options":json.loads(r[5])} for r in rows]
    finally:
        conn.close()

@app.get("/api/admin/participant/{emp_id}")
def get_participant_detail(emp_id: str, _=Depends(require_admin)):
    conn = get_conn()
    try:
        p = db_run(conn, "SELECT emp_id,name,rounds FROM participants WHERE emp_id=%s", emp_id)
        if not p: raise HTTPException(404,"Participant not found")
        preds = db_run(conn, "SELECT round,match_name,prediction FROM predictions WHERE emp_id=%s ORDER BY round", emp_id)
        pts = db_run(conn, "SELECT r1_pts,r2_pts,r3_pts,COALESCE(bonus_pts,0),total FROM points_cache WHERE emp_id=%s", emp_id)
        match_res = db_run(conn, "SELECT round,match_name,result FROM matches WHERE status='done'")
        mr = {(r[0],r[1]):r[2] for r in match_res}
        pred_detail = []
        for pr in preds:
            res = mr.get((pr[0],pr[1]))
            correct = check_pred(pr[2], res) if res else None
            pred_detail.append({"round":pr[0],"match":pr[1],"prediction":pr[2],
                                 "result":res,"correct":correct,"points":3 if correct else 0})
        pts_row = pts[0] if pts else (0,0,0,0,0)
        return {"emp_id":emp_id,"name":p[0][1],"rounds":json.loads(p[0][2]),
                "r1_pts":float(pts_row[0]),"r2_pts":float(pts_row[1]),
                "r3_pts":float(pts_row[2]),"bonus_pts":float(pts_row[3]),
                "total":float(pts_row[4]),"predictions":pred_detail}
    finally:
        conn.close()

@app.post("/api/admin/reset-and-reseed")
def reset_reseed(_=Depends(require_admin)):
    try:
        conn = get_conn()
        try:
            print("Dropping tables...")
            db_run(conn, "DROP TABLE IF EXISTS points_cache CASCADE")
            db_run(conn, "DROP TABLE IF EXISTS predictions CASCADE")
            db_run(conn, "DROP TABLE IF EXISTS matches CASCADE")
            db_run(conn, "DROP TABLE IF EXISTS participants CASCADE")
            print("Tables dropped.")
            init_db(conn)
            print("Tables created, seeding...")
            seed_db(conn)
            count = db_run(conn, "SELECT COUNT(*) FROM participants")[0][0]
            matches = db_run(conn, "SELECT COUNT(*) FROM matches")[0][0]
            print(f"Done: {count} participants, {matches} matches.")
            return {"status":"ok","participants":count,"matches":matches}
        finally:
            conn.close()
    except Exception as e:
        traceback.print_exc()
        return {"status":"error","detail":str(e)}

@app.get("/api/admin/test-seed")
def test_seed(_=Depends(require_admin)):
    try:
        conn = get_conn()
        try:
            seed_path = os.path.join(os.path.dirname(__file__), "..", "data", "seed_data.json")
            with open(seed_path, encoding="utf-8") as f:
                data = json.load(f)
            participants = data["participants"]
            match_options = data.get("match_options", {})
            preds_count = sum(len(info["predictions"]) for info in participants.values())
            db_check = db_run(conn, "SELECT COUNT(*) FROM participants")[0][0]
            return {
                "status": "ok",
                "seed_participants": len(participants),
                "seed_predictions": preds_count,
                "seed_matches": sum(len(data["matches"][k]) for k in ("R1","R2","R3")),
                "match_options_count": len(match_options),
                "db_participants_now": db_check,
            }
        finally:
            conn.close()
    except Exception as e:
        traceback.print_exc()
        return {"status": "error", "detail": str(e)}


@app.post("/api/admin/merge-participant")
def merge_participant(payload: dict, _=Depends(require_admin)):
    """Merge wrong_id into correct_id. Preserves all match results and bonus points."""
    wrong_id  = str(payload.get("wrong_id",  "")).strip()
    correct_id = str(payload.get("correct_id", "")).strip()
    if not wrong_id or not correct_id:
        raise HTTPException(400, "wrong_id and correct_id are required")
    if wrong_id == correct_id:
        raise HTTPException(400, "IDs are the same")

    conn = get_conn()
    try:
        w = db_run(conn, "SELECT emp_id,name FROM participants WHERE emp_id=%s", wrong_id)
        c = db_run(conn, "SELECT emp_id,name FROM participants WHERE emp_id=%s", correct_id)
        if not w: raise HTTPException(404, f"wrong_id {wrong_id} not found")
        if not c: raise HTTPException(404, f"correct_id {correct_id} not found")

        wrong_name   = w[0][1]
        correct_name = c[0][1]

        # Move predictions not already covered by correct_id
        wrong_preds = db_run(conn,
            "SELECT round,match_name,prediction FROM predictions WHERE emp_id=%s", wrong_id)
        existing = set(
            (r[0], r[1]) for r in db_run(conn,
                "SELECT round,match_name FROM predictions WHERE emp_id=%s", correct_id))
        moved = 0
        for rnd, match, pred in wrong_preds:
            if (rnd, match) not in existing:
                safe_run(conn,
                    "INSERT INTO predictions (emp_id,round,match_name,prediction) VALUES (%s,%s,%s,%s) ON CONFLICT DO NOTHING",
                    correct_id, rnd, match, pred)
                moved += 1

        # Merge rounds list
        wr = db_run(conn, "SELECT rounds FROM participants WHERE emp_id=%s", wrong_id)
        cr = db_run(conn, "SELECT rounds FROM participants WHERE emp_id=%s", correct_id)
        w_rounds = json.loads(wr[0][0]) if wr else []
        c_rounds = json.loads(cr[0][0]) if cr else []
        merged = sorted(set(w_rounds + c_rounds))
        db_run(conn, "UPDATE participants SET rounds=%s WHERE emp_id=%s",
               json.dumps(merged), correct_id)

        # Delete wrong_id
        db_run(conn, "DELETE FROM predictions  WHERE emp_id=%s", wrong_id)
        db_run(conn, "DELETE FROM points_cache WHERE emp_id=%s", wrong_id)
        db_run(conn, "DELETE FROM participants  WHERE emp_id=%s", wrong_id)

        # Recalculate points for correct_id only
        preds = db_run(conn, "SELECT round,match_name,prediction FROM predictions WHERE emp_id=%s", correct_id)
        results_rows = db_run(conn, "SELECT round,match_name,result FROM matches WHERE status='done' AND result IS NOT NULL")
        results = {(r[0],r[1]): r[2] for r in results_rows}
        pts = {1:0.0, 2:0.0, 3:0.0}
        for pr in preds:
            key = (pr[0], pr[1])
            if key in results and pr[2] and check_pred(pr[2], results[key]):
                pts[pr[0]] += 3.0
        bonus_row = db_run(conn, "SELECT bonus_pts FROM points_cache WHERE emp_id=%s", correct_id)
        bonus_pts = float(bonus_row[0][0]) if bonus_row else 0.0
        total = pts[1] + pts[2] + pts[3] + bonus_pts
        safe_run(conn, """INSERT INTO points_cache (emp_id,r1_pts,r2_pts,r3_pts,bonus_pts,total)
            VALUES (%s,%s,%s,%s,%s,%s)
            ON CONFLICT (emp_id) DO UPDATE SET
            r1_pts=EXCLUDED.r1_pts, r2_pts=EXCLUDED.r2_pts,
            r3_pts=EXCLUDED.r3_pts, total=EXCLUDED.total""",
            correct_id, pts[1], pts[2], pts[3], bonus_pts, total)

        return {
            "status": "ok",
            "merged": f"{wrong_id} ({wrong_name}) → {correct_id} ({correct_name})",
            "predictions_moved": moved,
            "rounds": merged,
            "new_total_pts": total
        }
    finally:
        conn.close()


@app.post("/api/admin/rename-participant")
def rename_participant(payload: dict, _=Depends(require_admin)):
    """Change an employee ID to a new one (e.g. fix wrong ID)."""
    old_id = str(payload.get("old_id", "")).strip()
    new_id = str(payload.get("new_id", "")).strip()
    if not old_id or not new_id:
        raise HTTPException(400, "old_id and new_id are required")
    if old_id == new_id:
        raise HTTPException(400, "IDs are the same")
    conn = get_conn()
    try:
        p = db_run(conn, "SELECT emp_id,name FROM participants WHERE emp_id=%s", old_id)
        if not p: raise HTTPException(404, f"ID {old_id} not found")
        existing = db_run(conn, "SELECT emp_id FROM participants WHERE emp_id=%s", new_id)
        if existing: raise HTTPException(409, f"ID {new_id} already exists — use merge instead")
        # Rename across all tables
        db_run(conn, "INSERT INTO participants (emp_id,name,rounds) SELECT %s,name,rounds FROM participants WHERE emp_id=%s", new_id, old_id)
        db_run(conn, "UPDATE predictions SET emp_id=%s WHERE emp_id=%s", new_id, old_id)
        db_run(conn, "INSERT INTO points_cache (emp_id,r1_pts,r2_pts,r3_pts,bonus_pts,total) SELECT %s,r1_pts,r2_pts,r3_pts,COALESCE(bonus_pts,0),total FROM points_cache WHERE emp_id=%s", new_id, old_id)
        db_run(conn, "DELETE FROM points_cache WHERE emp_id=%s", old_id)
        db_run(conn, "DELETE FROM predictions  WHERE emp_id=%s", old_id)
        db_run(conn, "DELETE FROM participants  WHERE emp_id=%s", old_id)
        return {"status": "ok", "renamed": f"{old_id} → {new_id}", "name": p[0][1]}
    finally:
        conn.close()

@app.post("/api/admin/recalculate")
def recalculate_points(_=Depends(require_admin)):
    """Recalculate points for all participants based on current match results. Bonus points are preserved."""
    try:
        conn = get_conn()
        try:
            recalc_all_points(conn)
        finally:
            conn.close()
        return {"status": "ok", "message": "Points recalculated for all participants"}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))

@app.get("/api/admin/export/excel")
def export_excel(_=Depends(require_admin)):
    """Export full leaderboard as Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        conn = get_conn()
        try:
            rows = db_run(conn, """
                SELECT pc.emp_id, p.name, p.rounds,
                       pc.r1_pts, pc.r2_pts, pc.r3_pts,
                       COALESCE(pc.bonus_pts,0), pc.total
                FROM points_cache pc
                JOIN participants p ON pc.emp_id=p.emp_id
                ORDER BY pc.total DESC, p.name ASC
            """)
        finally:
            conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Leaderboard"

        # Styles
        def fill(hex6):
            return PatternFill('solid', fgColor='FF'+hex6)
        def hfont():
            return Font(name='Arial', bold=True, color='FFFFFFFF', size=10)
        def cfont(bold=False):
            return Font(name='Arial', bold=bold, size=10)
        def border():
            s = Side(style='thin', color='FFCCCCCC')
            return Border(left=s, right=s, top=s, bottom=s)

        headers = ['#', 'Employee ID', 'Name', 'Rounds', 'R1 Points', 'R2 Points', 'R3 Points', 'Bonus Points', 'Total']
        widths  = [5, 14, 32, 10, 11, 11, 11, 13, 10]

        # Header row
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = fill('1A1A1A')
            c.font = hfont()
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 20

        # Data rows
        for rank, row in enumerate(rows, 1):
            emp_id, name, rounds_json, r1, r2, r3, bonus, total = row
            rounds = json.loads(rounds_json) if rounds_json else []
            rounds_str = '+'.join([f'R{r}' for r in sorted(rounds)])
            er = rank + 1
            data = [rank, emp_id, name, rounds_str, float(r1), float(r2), float(r3), float(bonus), float(total)]
            medal = {1:'FFD700', 2:'C0C0C0', 3:'CD7F32'}
            for ci, val in enumerate(data, 1):
                c = ws.cell(row=er, column=ci, value=val)
                c.font = cfont(bold=(ci==9))
                c.border = border()
                c.alignment = Alignment(horizontal='center' if ci!=3 else 'left', vertical='center')
                if rank in medal:
                    c.fill = fill(medal[rank])
                elif er % 2 == 0:
                    c.fill = fill('F5F5F5')
                if ci == 9:
                    c.fill = fill('FCE4D6')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"SMI_FIFA2026_Leaderboard_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'}
        )
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/admin/export/pdf")
def export_pdf(_=Depends(require_admin)):
    """Export full leaderboard as PDF with proper Arabic text support."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import arabic_reshaper
        from bidi.algorithm import get_display

        # Register Arabic font
        font_path = os.path.join(os.path.dirname(__file__), "fonts", "Amiri-Regular.ttf")
        pdfmetrics.registerFont(TTFont('Amiri', font_path))

        def ar(text):
            """Reshape and reorder Arabic text for correct PDF rendering."""
            try:
                reshaped = arabic_reshaper.reshape(str(text))
                return get_display(reshaped)
            except:
                return str(text)

        conn = get_conn()
        try:
            rows = db_run(conn, """
                SELECT pc.emp_id, p.name, p.rounds,
                       pc.r1_pts, pc.r2_pts, pc.r3_pts,
                       COALESCE(pc.bonus_pts,0), pc.total
                FROM points_cache pc
                JOIN participants p ON pc.emp_id=p.emp_id
                ORDER BY pc.total DESC, p.name ASC
            """)
            done = db_run(conn, "SELECT COUNT(*) FROM matches WHERE status='done'")[0][0]
        finally:
            conn.close()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                rightMargin=15*mm, leftMargin=15*mm,
                                topMargin=15*mm, bottomMargin=15*mm)

        title_style = ParagraphStyle('title', fontSize=16, fontName='Helvetica-Bold',
                                     textColor=colors.HexColor('#1A1A1A'), spaceAfter=4)
        sub_style = ParagraphStyle('sub', fontSize=10, fontName='Helvetica',
                                   textColor=colors.HexColor('#666666'), spaceAfter=12)

        elements = []
        elements.append(Paragraph("SMI Mechanical — FIFA World Cup 2026", title_style))
        elements.append(Paragraph(
            f"Prediction League Leaderboard  |  Generated: {datetime.utcnow().strftime('%d %b %Y %H:%M')} UTC  |  Matches played: {done}/104",
            sub_style))
        elements.append(Spacer(1, 4*mm))

        col_headers = ['#', 'Emp ID', 'Name', 'Rounds', 'R1', 'R2', 'R3', 'Bonus', 'Total']
        table_data = [col_headers]
        for rank, row in enumerate(rows, 1):
            emp_id, name, rounds_json, r1, r2, r3, bonus, total = row
            rounds = json.loads(rounds_json) if rounds_json else []
            rounds_str = '+'.join([f'R{r}' for r in sorted(rounds)])
            # Apply Arabic reshaping to name
            display_name = ar(name)
            table_data.append([
                str(rank), str(emp_id), display_name, rounds_str,
                str(int(r1)), str(int(r2)), str(int(r3)),
                str(int(bonus)), str(int(total))
            ])

        col_widths = [12*mm, 20*mm, 72*mm, 20*mm, 16*mm, 16*mm, 16*mm, 18*mm, 18*mm]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)

        gold   = colors.HexColor('#FFD700')
        silver = colors.HexColor('#C0C0C0')
        bronze = colors.HexColor('#CD7F32')
        dark   = colors.HexColor('#1A1A1A')
        light  = colors.HexColor('#F5F5F5')
        peach  = colors.HexColor('#FCE4D6')

        style_cmds = [
            ('BACKGROUND', (0,0), (-1,0), dark),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,0), 9),
            ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME',   (2,1), (2,-1), 'Amiri'),   # Arabic font for names
            ('FONTSIZE',   (2,1), (2,-1), 9),
            ('ALIGN',      (2,1), (2,-1), 'RIGHT'),   # RTL alignment for names
            ('FONTSIZE',   (0,1), (-1,-1), 8),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light]),
            ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#CCCCCC')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('BACKGROUND', (8,1), (8,-1), peach),
            ('FONTNAME',   (8,1), (8,-1), 'Helvetica-Bold'),
        ]
        if len(table_data) > 1: style_cmds.append(('BACKGROUND', (0,1), (-1,1), gold))
        if len(table_data) > 2: style_cmds.append(('BACKGROUND', (0,2), (-1,2), silver))
        if len(table_data) > 3: style_cmds.append(('BACKGROUND', (0,3), (-1,3), bronze))

        t.setStyle(TableStyle(style_cmds))
        elements.append(t)
        doc.build(elements)
        buf.seek(0)

        fname = f"SMI_FIFA2026_Leaderboard_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.pdf"
        return StreamingResponse(buf, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'})
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))



@app.post("/api/admin/add-round4")
async def add_round4(_=Depends(require_admin)):
    """Batch add Round of 32 in background thread to avoid timeout."""
    import threading
    seed_path = os.path.join(os.path.dirname(__file__), "..", "data", "r4_data.json")
    if not os.path.exists(seed_path):
        raise HTTPException(404, "r4_data.json not found in data/ folder")

    with open(seed_path, encoding="utf-8") as f:
        r4 = json.load(f)

    def do_add():
        try:
            conn = get_conn()
            stats = {"new_participants": 0, "predictions_added": 0, "matches_added": 0}
            print("Starting R4 migration...")

            # Step 1: Get existing participants and their rounds in one query
            existing_rows = conn.run("SELECT emp_id, rounds FROM participants")
            existing = {r[0]: json.loads(r[1]) for r in existing_rows}

            # Step 2: Add new participants in one batch
            new_emps = [(eid, info["name"]) for eid, info in r4["participants"].items() if eid not in existing]
            if new_emps:
                ph = ",".join(f"(${i*3+1},${i*3+2},'[4]')" for i in range(len(new_emps)))
                flat = [v for row in new_emps for v in row]
                conn.run(f"INSERT INTO participants (emp_id,name,rounds) VALUES {ph} ON CONFLICT DO NOTHING", *flat)
                # Add points_cache rows
                ph2 = ",".join(f"(${i+1},0,0,0,0,0)" for i in range(len(new_emps)))
                flat2 = [r[0] for r in new_emps]
                conn.run(f"INSERT INTO points_cache (emp_id,r1_pts,r2_pts,r3_pts,bonus_pts,total) VALUES {ph2} ON CONFLICT DO NOTHING", *flat2)
                stats["new_participants"] = len(new_emps)
                print(f"Added {len(new_emps)} new participants")

            # Step 3: Update rounds for existing participants
            for eid, info in r4["participants"].items():
                if eid in existing and 4 not in existing[eid]:
                    new_rounds = json.dumps(sorted(existing[eid] + [4]))
                    conn.run("UPDATE participants SET rounds=$1 WHERE emp_id=$2", new_rounds, eid)

            print("Updated rounds")

            # Step 4: Get existing R4 predictions to skip
            existing_preds = set()
            existing_pred_rows = conn.run("SELECT emp_id, match_name FROM predictions WHERE round=4")
            for r in existing_pred_rows:
                existing_preds.add((r[0], r[1]))

            # Step 5: Batch insert predictions in chunks of 200
            all_preds = []
            for eid, info in r4["participants"].items():
                for mn, pv in info["predictions"].items():
                    if (eid, mn) not in existing_preds:
                        all_preds.append((eid, mn, pv or ""))

            chunk = 200
            for i in range(0, len(all_preds), chunk):
                c = all_preds[i:i+chunk]
                ph = ",".join(f"(${j*3+1},4,${j*3+2},${j*3+3})" for j in range(len(c)))
                flat = [v for row in c for v in row]
                conn.run(f"INSERT INTO predictions (emp_id,round,match_name,prediction) VALUES {ph} ON CONFLICT DO NOTHING", *flat)
                print(f"Predictions chunk {i//chunk+1} done ({len(c)} rows)")

            stats["predictions_added"] = len(all_preds)

            # Step 6: Add matches
            existing_matches = set(r[0] for r in conn.run("SELECT match_name FROM matches WHERE round=4"))
            match_rows = [(mn, json.dumps(r4["options"].get(mn,[]), ensure_ascii=False))
                          for mn in r4["matches"] if mn not in existing_matches]
            if match_rows:
                ph = ",".join(f"(4,${i*2+1},'pending',${i*2+2})" for i in range(len(match_rows)))
                flat = [v for row in match_rows for v in row]
                conn.run(f"INSERT INTO matches (round,match_name,status,options) VALUES {ph} ON CONFLICT DO NOTHING", *flat)
                stats["matches_added"] = len(match_rows)
                print(f"Added {len(match_rows)} matches")

            # Step 7: Add r4_pts column if needed
            try:
                conn.run("ALTER TABLE points_cache ADD COLUMN IF NOT EXISTS r4_pts REAL DEFAULT 0")
            except Exception:
                pass

            conn.run("UPDATE points_cache SET total=COALESCE(r1_pts,0)+COALESCE(r2_pts,0)+COALESCE(r3_pts,0)+COALESCE(r4_pts,0)+COALESCE(bonus_pts,0)")
            conn.close()
            print(f"R4 migration complete: {stats}")
        except Exception as e:
            traceback.print_exc()
            print(f"R4 migration FAILED: {e}")

    threading.Thread(target=do_add, daemon=True).start()
    return {"status": "ok", "message": "R4 migration started — check /api/health in 30s for participant count"}

# ── Serve frontend ────────────────────────────────────────
frontend_path = os.path.join(os.path.dirname(__file__), "..", "frontend")
if os.path.exists(frontend_path):
    app.mount("/", StaticFiles(directory=frontend_path, html=True), name="frontend")
